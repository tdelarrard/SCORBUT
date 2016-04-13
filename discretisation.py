"""Lire les données de discretisation du problème.

Ce module lire les données d'un fichier d'Excel qui discretise le problème de 
séchage.

Attributes
----------
- `L`             -- (float)   | Longueur [m].
- `J`             -- (int)     | Nombre de divisions [-].
- `T`             -- (ndarray) | Temps d'analyse [s].
- `N`             -- (int)     | Nombre de pas d'analyse [-].
- `erreur_max`    -- (float)   | Erreur maximum absolu relative [-].
- `h_all_0`       -- (ndarray) | Humidité relative initiale [-].
- `h_all_1_0`     -- (ndarray) | Humidité relative supposée pour le premier pas [-].
- `condLim`       -- (ndarray) | Type de conditions limites [-].
- `condLim_0_all` -- (ndarray) | Conditions limites du nœud à gauche [-] ou [1/m].
- `condLim_L_all` -- (ndarray) | Conditions limites du nœud à droite [-] ou [1/m].
- `temp`          -- (ndarray) | Température [K].
- `p_l`           -- (ndarray) | Masse volumique de l'eau [kg/m³].
- `n_l`           -- (ndarray) | Viscosité dynamique de l'eau [Pa·s].
- `Pvsat`         -- (ndarray) | Pression de vapeur saturante [Pa].
- `M_l`           -- (float)   | Masse molaire de l'eau [kg/mol].
- `R`             -- (float)   | Constante des gaz parfaits [J/mol·K].
- `phi`           -- (float)   | Porosité [-].
- `Sr`            -- (ndarray) | Coefficients du polynôme du taux de la saturation 
                                 du béton [-].
- `k_l`           -- (float)   | Perméabilité à l'eau liquide [m²].
- `m`             -- (float)   | Paramètre de Van Genuchten [-].
- `D_0`           -- (float)   | Coefficient de diffusion de la vapeur d'eau hors 
                                 milieu poreux [m²/s].
- `A`             -- (float)   | Coefficient A [-].
- `B`             -- (float)   | Coefficient B [-].

Notes
-----
1. - L'humidité relative supposée pour le premier pas `h_all_1_0` est égale à 
     l'humidité relative initiale `h_all_0` par défaut.
2. - Les types de conditions limites `condLim` sont supposées de Dirichlet dans le 
     nœud le plus à gauche et de Neumann dans le nœud le plus à droite.
3. - Les valeurs des conditions limites du nœud à droite `condLim_L_all` sont 
     supposées égales à 0 [m⁻¹].
4. - La masse volumique de l'eau `p_l`, la viscosité dynamique de l'eau `n_l` et la 
     pression de vapeur saturante `Pvsat` sont calculés en fonction de la 
     temperature `temp`.
5. - La valeur de la masse molaire de l'eau `M_l` et la constante des gaz parfaits 
     `R` sont definies dans le module `constantes`.
6. - Possibilité de prendre en compte la variation de la porosité dans l'espace (?).

"""
from fonctions  import masseVolumiqueLEau, presionVapeurSaturante, table, \
                       viscositeDynamique
from openpyxl   import load_workbook
from numpy      import array, concatenate, empty, linspace, shape
from constantes import codenames_sheets, livreExcel, M_l, noms_donnees, R

import os

__all__ = ['L', 'J', 'T', 'N', 'erreur_max', 'h_all_0', 'h_all_1_0', 'condLim', 
           'condLim_0_all', 'condLim_L_all', 'temp', 'p_l', 'n_l', 'Pvsat', 'M_l', 
           'R', 'phi', 'Sr', 'k_l', 'm', 'D_0', 'A', 'B']

# Declaration de variables
ts = os.get_terminal_size() # : dimensions de la terminal
larguer_terminal = ts.columns   # : larguer de la terminal

# Début : mettre les données du problème #

# Charger données du fichier Données.xlsm
strTexte = "Chargement les données du fichier"
print(strTexte.center(larguer_terminal,'*'))

# Ouvrir le livre Données.xlsm
wb = load_workbook(filename  = livreExcel, # livre d'Excel à ouvrir    
                   read_only = True,       # seulement lire, on peut pas modifier
                   data_only = True)       # lire les valeurs des cellules

# Mettre les données

# Discretisation du problème
strTexte = codenames_sheets['Discretisation']
print(strTexte.center(larguer_terminal, '-'))

try:
    # Ovrir la feuille Discretisation
    ws = wb[codenames_sheets['Discretisation']]
    discProbl = []
    for i in range(5):
        try:
            discProbl += [ws['D' + str(4 + i)].value]
            print("{} : {:9.3e}".format(noms_donnees[i], discProbl[i]))
        except ValueError:
            discProbl += [eval(ws['D'+str(4 + i)].value.strip('='))]
            print("{} : {:9.3e}".format(noms_donnees[i], discProbl[i]))
    
    L, J, T, N, erreur_max = discProbl
    N, J = int(N), int(J)
    
except:
    strTexte = "Erreur des données d'entrée. Veuillez de vérifier les " + \
               "données (" + noms_donnees[i] + ") et  puis faire pression " + \
               "sur Calculer autre fois. \n"
    input(strTexte)
    exit()

# Conditions initiales
strTexte = codenames_sheets['CondsInitiales']
print('\n', strTexte.center(larguer_terminal, '-'), sep = '')

try:
    # Ouvrir la feuille CondsInitiales
    ws = wb[codenames_sheets['CondsInitiales']]
    h_all_0 = empty([J + 1])
    
    for i in range(int(J) + 1):
        h_all_0[i] = ws['C' + str(4 + i)].value

    formate = ['{:6.4f}', '{:9.3e}']
    strTexte = ['x [m]', 'Humidité Iniciale [-]']
    coords = linspace(0, L, J + 1)
    condsInitiales = array([coords, h_all_0]).T
    listTable = table(strTexte, condsInitiales, formate, ' ', larguer_terminal)
    print(*listTable, sep = '')
    
except:
    strTexte = "Erreur des données d'entrée. Veuillez de vérifier les " + \
               "données (cellule C" + str(4 + i) + "), et  puis faire " + \
               "pression sur Calculer autre fois. \n"
    input(strTexte)
    exit()

h_all_1_0 = h_all_0

# Conditions limites
strTexte = codenames_sheets['CondsLimites']
print(strTexte.center(larguer_terminal, '-'))

try:
    # Ouvrir la feuille CondsLimites
    ws = wb[codenames_sheets['CondsLimites']]
    T = empty([N + 1])
    condLim_0_all = empty([N + 1])
    
    for i in range(int(N) + 1):
        T[i] = ws['B' + str(4 + i)].value
        condLim_0_all[i] = ws['C' + str(4 + i)].value
    
    formate = ['{:012.3f}', '{:9.3e}']
    strTexte = ['t [s]', 'Conditions Limites à Gauche [-]']
    condsLimites = array([T, condLim_0_all]).T
    listTable = table(strTexte, condsLimites, formate, ' ', larguer_terminal)
    print(*listTable, sep = '')
    
except:
    strTexte = "Erreur des données d'entrée. Veuillez de vérifier les " + \
               "données (cellule C" + str(4 + i) + "), et  puis faire " + \
               "pression sur Calculer autre fois. \n"
    input(strTexte)
    exit()

condLim = array([0, 1])
condLim_L_all = 0 * condLim_0_all

# Température
strTexte = codenames_sheets['Temperature']
print(strTexte.center(larguer_terminal, '-'))

try:
    # Ouvrir la feuille Temperature
    ws = wb[codenames_sheets['Temperature']]
    temp = empty([N + 1])
    
    for i in range(int(N) + 1):
        temp[i] = ws['C' + str(4 + i)].value
    
    formate = ['{:012.3f}', '{:5.1f}']
    strTexte = ['t [s]', 'Temperature [K]']
    temperature = array([T, temp]).T
    listTable = table(strTexte, temperature, formate, ' ', larguer_terminal)
    print(*listTable, sep = '')
    
except:
    strTexte = "Erreur des données d'entrée. Veuillez de vérifier les " + \
               "données (cellule C" + str(4 + i) + "), et  puis faire " + \
               "pression sur Calculer autre fois. \n"
    input(strTexte)
    exit()

# Masse volumique
strTexto = "Masse Volumique"
print(strTexto.center(larguer_terminal, '-'))
p_l = masseVolumiqueLEau(temp)
formate = ['{:012.3f}', '{:7.3f}']
strTexte = ['t [s]', 'Masse Volumique [kg/m³]']
masseVolumique = array([T, p_l]).T
listTable = table(strTexte, masseVolumique, formate, ' ', larguer_terminal)
print(*listTable, sep = '')

# Viscosité dynamique de l'eau
strTexto = "Viscosité dynamique de l'eau"
print(strTexto.center(larguer_terminal, '-'))
n_l = viscositeDynamique(temp)
formate = ['{:012.3f}', '{:9.3e}']
strTexte = ['t [s]', 'Viscosité dynamique [Pa·s]']
viscositeDynamique = array([T, n_l]).T
listTable = table(strTexte, viscositeDynamique, formate, ' ', larguer_terminal)
print(*listTable, sep = '')

# Presion de vapeur saturante
strTexto = "Pression de Vapeur Saturante"
print(strTexto.center(larguer_terminal, '-'))
Pvsat = presionVapeurSaturante(temp)
formate = ['{:012.3f}', '{:8.3f}']
strTexte = ['t [s]', 'Pression de Vapeur Saturante [Pa]']
presionVapeurSaturante = array([T, Pvsat]).T
listTable = table(strTexte, presionVapeurSaturante, formate, ' ', larguer_terminal)
print(*listTable, sep = '')

# Propiétés du béton
strTexto = "Propiétés du Béton"
print(strTexto.center(larguer_terminal, '-'))
noms_propBeton = [noms_donnees[16], noms_donnees[17], noms_donnees[18],
                  noms_donnees[19], noms_donnees[20], noms_donnees[21],
                  noms_donnees[22]]

try:
    # Ovrir la feuille Discretisation
    ws = wb[codenames_sheets["PropBeton"]]
    propBeton = []
    for i in range(7):
        try:
            propBeton += [float(ws['D' + str(4+i)].value)]
            print('{} : {:9.3e}'.format(noms_propBeton[i], propBeton[i]))
        except:
            propBeton = propBeton + [eval("array(" + ws['D'+str(4+i)].value + ')')]
            strTexto = "{0} : {1}"
            print(strTexto.format(noms_propBeton[i], propBeton[i]))
        
    phi, Sr, k_l, m, D_0, A, B = propBeton

except:
    strTexto = "Erreur des données d'entrée. Veuillez de vérifier les " + \
               "données (" + noms_propBeton[i] + ") et  puis faire pression " + \
               "sur Calculer autre fois. \n"
    input(strTexto)
    exit()

strTexto = "Données d'entrée ont été bien mettre"
print('\n', strTexto.center(larguer_terminal, '+'), sep = '')
strTexto = "Chargement les données du fichier"
print(strTexto.center(larguer_terminal, '*'))
